import openpyxl

if __name__ == '__main__':
    wb = openpyxl.load_workbook('excel_file.xlsx')
    ws = wb.active
    for cell in ws['A']:
        print(cell.value)
    for cell in ws['B']:
        print(cell.value)
